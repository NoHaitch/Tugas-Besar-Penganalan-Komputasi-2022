# PROGRAM SIMULASI PINTU TOL BERBASIS KARTU E-MONEY
# membuat simulasi pintu tol yang berbasis kartu e-money dengan tujuan memahami struktur proses yang dilakukan pintu tol otomatis

"""
PEMBUAT:
    Kelompok 10
    -> 16522098 - Muhammad Farhan
    -> 16522208 - Hafidz Yasifa Patty
    -> 19622078 - Raden Francisco Trianto Bratadiningrat
    -> 19622188 - Mattheuw Suciadi Wijaya

ASUMSI :
    - input kartu digital yang seharusnya discan disimulasikan dengan menginput secara 
    manual data (kode kartu, saldo) yang dianggap jika format kode kartu sudah valid maka 
    anggap selalu benar tanpa perlu diverifikasi kodenya
    - anggap kendaraan sudah masuk tol dan sedang berada di pintu keluar tol sehingga langsung terkena biaya tol
KAMUS :
    kode_kartu, bantuan, selesai, path : string
    saldo, harga, saldoakhir : int
    simulasi, palang_pintu_gerbang, pembayaran, masukkankartu, input_akhir, input_tombol : bool
    input_kendaraan, input_gerbang : array string
    file : open("log.txt","a")

"""

# ALGORITMA

# menggunakan import
import datetime
import openpyxl as px


# Inisiasi Path Lokasi File Excel
path = 'Database_Card.xlsx'

# insialisasi fungsi
def daftar_kartu(a):
    # Mendaftarkan kartu dan saldonya ke dalam excel
    # ALGORITMA
    nomer = a
    saldo = inpsaldo()
    book = px.load_workbook(path)
    sheet = book.active
    baris_max = sheet.max_row
    nomer_id = sheet.cell(row=baris_max+1, column = 1)
    nomer_id.value = nomer
    saldo_id = sheet.cell(row=baris_max+1, column = 2)
    saldo_id.value = saldo
    book.save(path)

def bacasaldo(a):
    # mendapatkan saldo dari kartu yang sudah terdaftar di excel
    # ALGORITMA
    user_id = str(a)
    indeks = 0
    found = False
    while found == False:
        # Membuka file excel sebagai book
        book = px.load_workbook(path)
        sheet = book.active
        baris_max = sheet.max_row
        # Mencari apakah kode_kartu ada atau tidak
        for i in range(1, baris_max + 1):
            data_id = sheet.cell(row = i, column=1)
            if(user_id == str(data_id.value)):
                indeks = i
                found = True
                print("------------------Kartu Terdaftar----------------------")
        if(i == baris_max and found == False):
            print("Nomer tidak terdaftar, daftarkan kartu anda")
            print("-------------------------------------------")
            daftar_kartu(user_id)
    saldo = sheet.cell(row=indeks, column=2).value
    return saldo

def updatesaldo(a, b):
    # Mengubah data saldo di excel
    # ALGORITMA
    user_id = str(a)
    saldo_akhir = b
    # membuka file excel sebagai book
    book = px.load_workbook(path)
    sheet = book.active
    baris_max = sheet.max_row

    for i in range(1, baris_max + 1):
        data_id = sheet.cell(row = i, column=1)
        if(user_id == str(data_id.value)):
            indeks = i
    # merubah nilai saldo di kolom 2 dan baris tergantung ditemukannya kode_kartu yang sama
    saldo_id = sheet.cell(row=indeks, column = 2)
    saldo_id.value = saldo_akhir

    book.save(path)    

def inpkendaraan():
    # Mendapatkan golongan kendaraan yang valid
    # ALGORITMA
    input_kendaraan = False
    # pengulangan validasi input golongan kendaraan
    while input_kendaraan == False:
        print("Pilih golongan kendaraan:")
        print("~ golongan 1 (Rp5.000) -> sedan, jip, truk kecil, dan bus")
        print("~ golongan 2 (Rp8.000) -> truk dengan 2 gandar")
        print("~ golongan 3 (Rp11.000)-> truk dengan 3 gandar")
        print("Masukkan golongan kendaraan: ")
        gol_kendaraan = input(">>> ")
        if(gol_kendaraan.isdigit()):
            gol_kendaraan = int(gol_kendaraan)
            if(any(gol_kendaraan==num for num in [1,2,3])):
                input_kendaraan = True
                biaya_kendaraan = 5000+(gol_kendaraan-1)*3000
            else:
                print("Input golongan kendaran hanya antara 1 / 2 / 3")
        else:
            print("Input dalam bentuk angka saja")
    return [str(biaya_kendaraan), gol_kendaraan]

def inpkodekartu():
    # Mendapatkan kodekartu yang valid
    # ALGORITMA
    input_kodekartu = False
    # pengulangan validasi kode kartu
    while input_kodekartu == False:
        print("Masukan kode kartu: XXXX-XXXX-XXXX-XXXX / XXXXXXXXXXXXXXXX")
        kode_kartu = input(">>> ")
        if(len(kode_kartu) == 19):
            for digit_4 in kode_kartu.split("-"):
                if(not digit_4.isdigit()):
                    print("Kode kartu salah, contoh format yang benar XXXX-XXXX-XXXX-XXXX / XXXXXXXXXXXXXXXX")
                    break
            else: # Hasil Kode kartu berbentuk XXXX-XXXX-XXXX-XXXX
                input_kodekartu = True  
        elif(len(kode_kartu) == 16): 
            kode_benar = True
            for digit in kode_kartu:
                if(digit.isdigit() == False):
                    kode_benar = False
                    break
            else: # Hasil Kode kartu berbentuk XXXXXXXXXXXXXXXX
                input_kodekartu = True
                # Mengubah format kode kartu dari XXXXXXXXXXXXXXXX menjadi XXXX-XXXX-XXXX-XXXX
                temp_kode_kartu = kode_kartu
                kode_kartu = ""
                for i in range(4):
                    if(kode_kartu != ""):
                        kode_kartu += "-"
                    kode_kartu += temp_kode_kartu[(i*4):(i*4)+4]
            if(kode_benar == False):
                print("Kode kartu salah, contoh format yang benar XXXX-XXXX-XXXX-XXXX / XXXXXXXXXXXXXXXX")
        else:
            print("Kode kartu salah, contoh format yang benar XXXX-XXXX-XXXX-XXXX / XXXXXXXXXXXXXXXX")
    print(kode_kartu)
    return kode_kartu

def inpsaldo():
    # Mendapatkan saldo kartu yang valid
    # ALGORITMA
    input_saldo = False
    while input_saldo == False:
        print("Masukkan saldo kartu(Rp): ")
        saldo = input(">>> ")
        try:
            saldo = float(saldo)
        except:
            print("Format saldo salah, titik hanya sebagai penanda desimal")
        else:
            input_saldo = True
    return saldo

def inpgerbang():
    # mendapatkan biaya gerbang, gerbang masuk, dan gerbang keluar yang valid
    # ALGORITMA
        # Menampilkan data biaya tambahan pada setiap gerbang
        # Dibuat secara hardcoded yang dapat dilihat sebagai tabel berikut:
    print("Tabel harga dari gerbang ke gerbang dalam ribu rupiah (tidak termasuk biaya golongan kendaraan)")
    print(" | A | B | C | D | E ")
    print("A|X  |10 |14 |21 |26 ")
    print("B|10 |X  |12 |18 |22 ")
    print("C|14 |12 |X  |11 |16 ")
    print("D|21 |18 |11 |X  |21 ")    
    print("E|26 |22 |16 |21 |X ") # X diganti dengan nilai 0 (kosong) dalam matriksnya
    biaya = [[0, 10, 14, 21, 26], [10, 0, 12, 18, 22], [14, 12, 0, 11, 16], [21, 18, 11, 0,  21], [26, 22, 16, 21, 0]] 
    input_gerbangmasuk = False
    # pengulangan validasi input gerbang masuk
    while input_gerbangmasuk == False:
        masuk = input("Masukkan gerbang masuk: ")
        if(any(masuk == cek for cek in ["A","B","C","D","E"])):
            input_gerbangmasuk = True
        else:
            print("Masukkan gerbang hanya diantara: A/B/C/D/E")
    input_gerbangkeluar = False
    # pengulangan validasi input gerbang keluar
    while input_gerbangkeluar == False:
        keluar = input("Masukkan gerbang keluar: ")
        if(any(keluar == cek for cek in ["A","B","C","D","E"])):
            if(masuk == keluar):
                print("Gerbang masuk tidak boleh sama dengan gerbang keluar")
            else:
                input_gerbangkeluar = True
        else:
            print("Masukkan gerbang hanya diantara: A/B/C/D/E")
    biaya_gerbang = biaya[ord(masuk)-65][ord(keluar)-65]*1000
    return [str(biaya_gerbang),masuk,keluar]

# PEMBUKA PROGRAM
print("\n")
print("".center(60,"="))
print("Simulasi Gerbang Jalan Tol".center(60))

# LOOP SIMULASI
simulasi = True
while simulasi:
    # palang pintu gerbang, false jika tertutup, true jika sudah terbuka
    palang_pintu_gerbang = False

    # Input kendaraan = [golongan kendaraan, biaya dari golongan kendaraan]
    input_kendaraan = inpkendaraan()
    print()

    # Input kode kartu (16 angka string)
    kode_kartu = inpkodekartu()
    print()

    # Mendapat saldo kartu dari fungsi bacasaldo: mendapatkan saldo dari kode_kartu awal di excel, jika kode kartu tidak terdaftar maka meminta input saldo baru
    saldo = int(bacasaldo(kode_kartu))
    print()

    # Input Gerbang = [biaya_gerbang, masuk, keluar]
    input_gerbang = inpgerbang()
    print()

    # Mendapat harga total dari harga gerbang dan golongan kendaraan
    harga = (int(input_gerbang[0]) + int(input_kendaraan[0]))
    pembayaran = False
    print(f"Biaya total gerbang: {harga}")
    saldoakhir = saldo - harga
    # jika tidak bayar ( saldo akhir menjadi negatif )
    if(saldoakhir < 0):
        print("Pembayaran gagal, saldo tidak cukup")
        print("Masukkan kartu lainnya atau pencet tombol bantuan\n")            
        # apakah tombol bantuan ditekan atau tidak 
        input_tombol = False
        masukankartu = False
        while input_tombol == False:
            print("Tekan tombol bantuan? [ya/tidak] ")
            bantuan = input(">>> ")
            if(any(bantuan==char for char in ["y","Y","ya","Ya"])):
                print("Bantuan telah dipanggil")
                input_tombol = True
            elif(any(bantuan==char for char in ["t","T","tidak","Tidak"])):
                print("Bantuan tidak dipanggil")
                masukankartu = True
                input_tombol = True
            else:
                print("input antara y(iya) atau t(tidak)")
        if(masukankartu == True):
            # kembali ke input_kendaraan
            continue
    else: # jika bisa bayar
        pembayaran = True
        print("Pembayaran telah berhasil, palang telah dibuka")
        print(f"Sisa saldo Anda: {saldoakhir:.2f} rupiah")
        updatesaldo(kode_kartu, saldoakhir)
        palang_pintu_gerbang = True
        # mencatat di log.txt sebagai sejarah
        with open("log.txt","a") as file:
            file.write(f"\n{datetime.datetime.now().strftime('%c')}: {kode_kartu}; {saldo}; {saldoakhir}; {input_kendaraan[1]}; {input_gerbang[1]}; {input_gerbang[2]}")
        
    print()
    # apakah simulasi dilanjutkan atau tidak 
    input_akhir = False
    while input_akhir == False:
        print("lanjutkan simulasi? [ya/tidak]")
        selesai = input(">>> ")
        if(any(selesai==char for char in ["y","Y","ya","Ya"])):
            input_akhir = True
        elif(any(selesai==char for char in ["t","T","tidak","Tidak"])):
            print("Simulasi berakhir")
            input_akhir = True
            simulasi = False
            # Program Berhenti
        else:
            print("input antara y(iya) atau t(tidak)")